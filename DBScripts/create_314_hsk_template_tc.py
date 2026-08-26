from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def u(text: str) -> str:
    return text.encode("ascii").decode("unicode_escape")


HEADERS = [
    "GroupNo",
    "QuesNo",
    "Level",
    "Section",
    "PartCode",
    "GroupType",
    "GroupTitle",
    "InstructionText",
    "SharedPassage",
    "SharedOptionPool",
    "QuestionType",
    "Question",
    "QuestionImage",
    "Answer",
    "AnswerImage",
    "TrueAns",
    "QuestionOrder",
    "AnswerOrder",
    "Remark",
]


SHEETS = [
    {
        "name": "01_SharedWordBank",
        "title": u(r"314 HSK \u95b1\u8b80\u984c\u5eab\u532f\u5165\u7bc4\u672c"),
        "subtitle": u(r"\u984c\u578b\uff1a\u5171\u4eab\u8a5e\u8a9e\uff08SharedWordBank\uff09"),
        "note": u(r"\u9069\u7528\uff1aHSK4 ReadingPart1\u3002\u76f8\u540c GroupNo \u4ee3\u8868\u540c\u4e00\u7d44\u8a5e\u5eab\uff1b\u76f8\u540c QuesNo \u4ee3\u8868\u540c\u4e00\u984c\u3002"),
        "fill": "E2EFDA",
        "rows": [
            ["G001", "Q001", "HSK4", "Reading", "ReadingPart1", "SharedWordBank", u(r"HSK4 \u8a5e\u8a9e\u7d44 46-50"), u(r"\u7b2c 46-50 \u984c\uff1a\u9078\u8a5e\u586b\u7a7a\u3002"), "", u(r"A \u91cd  B \u9996\u5148  C \u89c0\u773e  D \u5805\u6301  E \u64e6  F \u5730\u9ede"), "SingleChoice", u(r"\u723a\u723a\uff0c\u70ba\u4ec0\u9ebc\u6a61\u76ae\u80fd\uff08 \uff09\u6389\u925b\u7b46\u5beb\u7684\u5b57\uff1f"), "", u(r"\u64e6"), "", 1, 46, 1, u(r"\u5171\u4eab\u8a5e\u8a9e\u7bc4\u4f8b")],
            ["G001", "Q001", "HSK4", "Reading", "ReadingPart1", "SharedWordBank", u(r"HSK4 \u8a5e\u8a9e\u7d44 46-50"), u(r"\u7b2c 46-50 \u984c\uff1a\u9078\u8a5e\u586b\u7a7a\u3002"), "", u(r"A \u91cd  B \u9996\u5148  C \u89c0\u773e  D \u5805\u6301  E \u64e6  F \u5730\u9ede"), "SingleChoice", u(r"\u723a\u723a\uff0c\u70ba\u4ec0\u9ebc\u6a61\u76ae\u80fd\uff08 \uff09\u6389\u925b\u7b46\u5beb\u7684\u5b57\uff1f"), "", u(r"\u89c0\u773e"), "", 0, 46, 2, u(r"\u5171\u4eab\u8a5e\u8a9e\u7bc4\u4f8b")],
            ["G001", "Q002", "HSK4", "Reading", "ReadingPart1", "SharedWordBank", u(r"HSK4 \u8a5e\u8a9e\u7d44 46-50"), u(r"\u7b2c 46-50 \u984c\uff1a\u9078\u8a5e\u586b\u7a7a\u3002"), "", u(r"A \u91cd  B \u9996\u5148  C \u89c0\u773e  D \u5805\u6301  E \u64e6  F \u5730\u9ede"), "SingleChoice", u(r"\u9019\u90e8\u96fb\u5f71\u975e\u5e38\u611f\u4eba\uff0c\u5f88\u591a\uff08 \uff09\u90fd\u88ab\u611f\u52d5\u5f97\u54ed\u4e86\u3002"), "", u(r"\u89c0\u773e"), "", 1, 47, 1, u(r"\u5171\u4eab\u8a5e\u8a9e\u7bc4\u4f8b")],
            ["G001", "Q002", "HSK4", "Reading", "ReadingPart1", "SharedWordBank", u(r"HSK4 \u8a5e\u8a9e\u7d44 46-50"), u(r"\u7b2c 46-50 \u984c\uff1a\u9078\u8a5e\u586b\u7a7a\u3002"), "", u(r"A \u91cd  B \u9996\u5148  C \u89c0\u773e  D \u5805\u6301  E \u64e6  F \u5730\u9ede"), "SingleChoice", u(r"\u9019\u90e8\u96fb\u5f71\u975e\u5e38\u611f\u4eba\uff0c\u5f88\u591a\uff08 \uff09\u90fd\u88ab\u611f\u52d5\u5f97\u54ed\u4e86\u3002"), "", u(r"\u5730\u9ede"), "", 0, 47, 2, u(r"\u5171\u4eab\u8a5e\u8a9e\u7bc4\u4f8b")],
        ],
    },
    {
        "name": "02_SentenceOrder",
        "title": u(r"314 HSK \u95b1\u8b80\u984c\u5eab\u532f\u5165\u7bc4\u672c"),
        "subtitle": u(r"\u984c\u578b\uff1a\u8a9e\u5e8f\u6392\u5217\uff08SentenceOrder\uff09"),
        "note": u(r"\u9069\u7528\uff1aHSK4 ReadingPart2\u3002Question \u6b04\u4f4d\u53ef\u76f4\u63a5\u653e A/B/C \u53e5\u6bb5\uff0cAnswer \u6b04\u4f4d\u653e\u6b63\u78ba\u9806\u5e8f\u3002"),
        "fill": "FFF2CC",
        "rows": [
            ["G002", "Q003", "HSK4", "Reading", "ReadingPart2", "SentenceOrder", u(r"HSK4 \u8a9e\u5e8f 56"), u(r"\u7b2c 56-65 \u984c\uff1a\u6392\u5217\u9806\u5e8f\u3002"), "", "", "SentenceOrder", u(r"A \u610f\u601d\u662f\u5e0c\u671b\u670b\u53cb\u4e4b\u9593\u7684\u53cb\u597d\u95dc\u4fc2  B \u80fd\u5920\u4e00\u76f4\u7e7c\u7e8c\u4e0b\u53bb\uff0c\u8d8a\u4e45\u8d8a\u597d  C \u4eba\u5011\u5e38\u8aaa\u300c\u53cb\u8abc\u5730\u4e45\u5929\u9577\u300d"), "", "C-A-B", "", 1, 56, 1, u(r"\u8a9e\u5e8f\u6392\u5217\u7bc4\u4f8b")],
            ["G002", "Q003", "HSK4", "Reading", "ReadingPart2", "SentenceOrder", u(r"HSK4 \u8a9e\u5e8f 56"), u(r"\u7b2c 56-65 \u984c\uff1a\u6392\u5217\u9806\u5e8f\u3002"), "", "", "SentenceOrder", u(r"A \u610f\u601d\u662f\u5e0c\u671b\u670b\u53cb\u4e4b\u9593\u7684\u53cb\u597d\u95dc\u4fc2  B \u80fd\u5920\u4e00\u76f4\u7e7c\u7e8c\u4e0b\u53bb\uff0c\u8d8a\u4e45\u8d8a\u597d  C \u4eba\u5011\u5e38\u8aaa\u300c\u53cb\u8abc\u5730\u4e45\u5929\u9577\u300d"), "", "A-B-C", "", 0, 56, 2, u(r"\u8a9e\u5e8f\u6392\u5217\u7bc4\u4f8b")],
        ],
    },
    {
        "name": "03_PassageCloze",
        "title": u(r"314 HSK \u95b1\u8b80\u984c\u5eab\u532f\u5165\u7bc4\u672c"),
        "subtitle": u(r"\u984c\u578b\uff1a\u5b8c\u5f62\u586b\u7a7a\uff08PassageCloze\uff09"),
        "note": u(r"\u9069\u7528\uff1aHSK5 ReadingPart1\u3002SharedPassage \u653e\u6574\u6bb5\u77ed\u6587\uff0cQuestion \u653e\u6bcf\u5c0f\u984c\u53e5\u5b50\u3002"),
        "fill": "FCE4D6",
        "rows": [
            ["G003", "Q004", "HSK5", "Reading", "ReadingPart1", "PassageCloze", u(r"HSK5 \u5b8c\u5f62 46-48"), u(r"\u7b2c 46-48 \u984c\uff1a\u8acb\u9078\u51fa\u6b63\u78ba\u7b54\u6848\u3002"), u(r"\u6709\u4e00\u500b\u5e74\u8f15\u4eba\u5728\u4e00\u5bb6\u516c\u53f8\u505a\u5f97\u5f88\u51fa\u8272\uff0c\u4ed6\u70ba\u81ea\u5df1\u8a2d\u8a08\u4e86\u4e00\u500b\u7f8e\u597d\u7684\u672a\u4f86\uff0c\u5c0d 46 \u5145\u6eff\u4fe1\u5fc3\u3002"), "", "SingleChoice", u(r"46\uff0e\u5c0d\uff08 \uff09\u5145\u6eff\u4fe1\u5fc3\u3002"), "", u(r"\u672a\u4f86"), "", 1, 46, 1, u(r"\u5b8c\u5f62\u586b\u7a7a\u7bc4\u4f8b")],
            ["G003", "Q004", "HSK5", "Reading", "ReadingPart1", "PassageCloze", u(r"HSK5 \u5b8c\u5f62 46-48"), u(r"\u7b2c 46-48 \u984c\uff1a\u8acb\u9078\u51fa\u6b63\u78ba\u7b54\u6848\u3002"), u(r"\u6709\u4e00\u500b\u5e74\u8f15\u4eba\u5728\u4e00\u5bb6\u516c\u53f8\u505a\u5f97\u5f88\u51fa\u8272\uff0c\u4ed6\u70ba\u81ea\u5df1\u8a2d\u8a08\u4e86\u4e00\u500b\u7f8e\u597d\u7684\u672a\u4f86\uff0c\u5c0d 46 \u5145\u6eff\u4fe1\u5fc3\u3002"), "", "SingleChoice", u(r"46\uff0e\u5c0d\uff08 \uff09\u5145\u6eff\u4fe1\u5fc3\u3002"), "", u(r"\u5929\u6c23"), "", 0, 46, 2, u(r"\u5b8c\u5f62\u586b\u7a7a\u7bc4\u4f8b")],
        ],
    },
    {
        "name": "04_SharedPassage",
        "title": u(r"314 HSK \u95b1\u8b80\u984c\u5eab\u532f\u5165\u7bc4\u672c"),
        "subtitle": u(r"\u984c\u578b\uff1a\u77ed\u6587\u95b1\u8b80\uff08SharedPassage\uff09"),
        "note": u(r"\u9069\u7528\uff1aHSK5 ReadingPart3\u3002SharedPassage \u653e\u5171\u7528\u77ed\u6587\uff0c\u540c\u4e00 GroupNo \u53ef\u5c0d\u61c9\u591a\u984c\u3002"),
        "fill": "DEEAF6",
        "rows": [
            ["G004", "Q005", "HSK5", "Reading", "ReadingPart3", "SharedPassage", u(r"HSK5 \u95b1\u8b80 71-74"), u(r"\u7b2c 71-74 \u984c\uff1a\u8acb\u6839\u64da\u77ed\u6587\u9078\u51fa\u6b63\u78ba\u7b54\u6848\u3002"), u(r"\u4e00\u500b\u51ac\u5929\uff0c\u4e00\u500b\u4eba\u5e36\u8457\u7375\u72d7\u53bb\u6253\u7375\u3002\u90a3\u500b\u4eba\u4e00\u69cd\u64ca\u4e2d\u4e86\u4e00\u96bb\u5154\u5b50\u7684\u817f\uff0c\u53d7\u50b7\u7684\u5154\u5b50\u62fc\u547d\u5730\u8dd1\u3002"), "", "SingleChoice", u(r"\u5154\u5b50\u7684\u817f\u600e\u9ebc\u4e86\uff1f"), "", u(r"\u88ab\u69cd\u6253\u4e2d\u4e86"), "", 1, 71, 1, u(r"\u77ed\u6587\u95b1\u8b80\u7bc4\u4f8b")],
            ["G004", "Q005", "HSK5", "Reading", "ReadingPart3", "SharedPassage", u(r"HSK5 \u95b1\u8b80 71-74"), u(r"\u7b2c 71-74 \u984c\uff1a\u8acb\u6839\u64da\u77ed\u6587\u9078\u51fa\u6b63\u78ba\u7b54\u6848\u3002"), u(r"\u4e00\u500b\u51ac\u5929\uff0c\u4e00\u500b\u4eba\u5e36\u8457\u7375\u72d7\u53bb\u6253\u7375\u3002\u90a3\u500b\u4eba\u4e00\u69cd\u64ca\u4e2d\u4e86\u4e00\u96bb\u5154\u5b50\u7684\u817f\uff0c\u53d7\u50b7\u7684\u5154\u5b50\u62fc\u547d\u5730\u8dd1\u3002"), "", "SingleChoice", u(r"\u5154\u5b50\u7684\u817f\u600e\u9ebc\u4e86\uff1f"), "", u(r"\u6454\u50b7\u4e86"), "", 0, 71, 2, u(r"\u77ed\u6587\u95b1\u8b80\u7bc4\u4f8b")],
            ["G004", "Q006", "HSK5", "Reading", "ReadingPart3", "SharedPassage", u(r"HSK5 \u95b1\u8b80 71-74"), u(r"\u7b2c 71-74 \u984c\uff1a\u8acb\u6839\u64da\u77ed\u6587\u9078\u51fa\u6b63\u78ba\u7b54\u6848\u3002"), u(r"\u4e00\u500b\u51ac\u5929\uff0c\u4e00\u500b\u4eba\u5e36\u8457\u7375\u72d7\u53bb\u6253\u7375\u3002\u90a3\u500b\u4eba\u4e00\u69cd\u64ca\u4e2d\u4e86\u4e00\u96bb\u5154\u5b50\u7684\u817f\uff0c\u53d7\u50b7\u7684\u5154\u5b50\u62fc\u547d\u5730\u8dd1\u3002"), "", "SingleChoice", u(r"\u7375\u72d7\u70ba\u4ec0\u9ebc\u6c92\u8ffd\u4e0a\u5154\u5b50\uff1f"), "", u(r"\u5154\u5b50\u62fc\u547d\u5730\u8dd1"), "", 1, 72, 1, u(r"\u77ed\u6587\u95b1\u8b80\u7bc4\u4f8b")],
            ["G004", "Q006", "HSK5", "Reading", "ReadingPart3", "SharedPassage", u(r"HSK5 \u95b1\u8b80 71-74"), u(r"\u7b2c 71-74 \u984c\uff1a\u8acb\u6839\u64da\u77ed\u6587\u9078\u51fa\u6b63\u78ba\u7b54\u6848\u3002"), u(r"\u4e00\u500b\u51ac\u5929\uff0c\u4e00\u500b\u4eba\u5e36\u8457\u7375\u72d7\u53bb\u6253\u7375\u3002\u90a3\u500b\u4eba\u4e00\u69cd\u64ca\u4e2d\u4e86\u4e00\u96bb\u5154\u5b50\u7684\u817f\uff0c\u53d7\u50b7\u7684\u5154\u5b50\u62fc\u547d\u5730\u8dd1\u3002"), "", "SingleChoice", u(r"\u7375\u72d7\u70ba\u4ec0\u9ebc\u6c92\u8ffd\u4e0a\u5154\u5b50\uff1f"), "", u(r"\u56e0\u70ba\u5b83\u7761\u8457\u4e86"), "", 0, 72, 2, u(r"\u77ed\u6587\u95b1\u8b80\u7bc4\u4f8b")],
        ],
    },
]


def build() -> None:
    wb = Workbook()
    wb.remove(wb.active)
    thin = Side(style="thin", color="808080")

    for sheet in SHEETS:
        ws = wb.create_sheet(sheet["name"])
        info_lines = [
            sheet["title"],
            sheet["subtitle"],
            sheet["note"],
            u(r"\u6b04\u4f4d\u8aaa\u660e\uff1aTrueAns=1 \u4ee3\u8868\u6b63\u78ba\u7b54\u6848\uff1b\u82e5\u6709\u5716\u7247\uff0c\u8acb\u5c07\u5716\u7247\u653e\u5728 Excel \u540c\u5c64\u7684 Images \u8cc7\u6599\u593e\u3002"),
        ]

        for idx, text in enumerate(info_lines, start=1):
            ws.cell(idx, 1, text)
            ws.merge_cells(start_row=idx, start_column=1, end_row=idx, end_column=len(HEADERS))

        for col, header in enumerate(HEADERS, start=1):
            ws.cell(6, col, header)

        for row_idx, row in enumerate(sheet["rows"], start=7):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row_idx, col_idx, value)

        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF", size=14)
            cell.fill = PatternFill("solid", fgColor="1F4E79")

        for row in range(2, 5):
            for col in range(1, len(HEADERS) + 1):
                cell = ws.cell(row, col)
                cell.fill = PatternFill("solid", fgColor="DDEBF7")
                cell.alignment = Alignment(wrap_text=True)

        for col in range(1, len(HEADERS) + 1):
            cell = ws.cell(6, col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9E1F2")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        last_row = 6 + len(sheet["rows"])
        for row in range(6, last_row + 1):
            for col in range(1, len(HEADERS) + 1):
                cell = ws.cell(row, col)
                cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        for row in range(7, last_row + 1):
            for col in range(1, len(HEADERS) + 1):
                ws.cell(row, col).fill = PatternFill("solid", fgColor=sheet["fill"])

        widths = {
            "A": 12, "B": 12, "C": 10, "D": 12, "E": 16, "F": 18, "G": 24,
            "H": 28, "I": 55, "J": 36, "K": 16, "L": 42, "M": 16, "N": 24,
            "O": 16, "P": 10, "Q": 14, "R": 12, "S": 30,
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        ws.row_dimensions[1].height = 24
        ws.row_dimensions[2].height = 22
        ws.row_dimensions[3].height = 22
        ws.row_dimensions[4].height = 36
        ws.freeze_panes = "A7"

    output = Path(__file__).with_name("314_HSK_Reading_Import_Template_4Sheets_TraditionalChinese.xlsx")
    wb.save(output)
    print(output)


if __name__ == "__main__":
    build()
