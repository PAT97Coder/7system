from pathlib import Path
import re
import shutil
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

SOURCE = Path(r"E:\01. Softwares Programming\24. Knowledge System\03. Documents\307\cau hoi chuyen nganh")
OUTPUT = SOURCE / "307_Upload"
HEADERS = [
    "QuestionId",
    "Question",
    "QuestionImage",
    "Reserved",
    "Answer",
    "AnswerImage",
    "IsCorrect",
]


def clean(value):
    if value is None:
        return ""
    return re.sub(r"[ \t\r\f\v]+", " ", str(value)).strip()


if OUTPUT.exists():
    resolved_output = OUTPUT.resolve()
    resolved_source = SOURCE.resolve()
    if resolved_output.parent != resolved_source:
        raise RuntimeError(f"Unsafe output path: {resolved_output}")
    shutil.rmtree(resolved_output)

OUTPUT.mkdir()

summary = []
for source_path in sorted(SOURCE.glob("*.xlsx")):
    source_book = openpyxl.load_workbook(source_path, data_only=True)
    if "EXCEL" not in source_book.sheetnames:
        raise RuntimeError(f"{source_path.name}: missing EXCEL sheet")
    source_sheet = source_book["EXCEL"]

    output_book = openpyxl.Workbook()
    output_sheet = output_book.active
    output_sheet.title = "EXCEL"
    output_sheet.append(HEADERS)

    question_count = 0
    answer_count = 0
    for source_row in range(2, source_sheet.max_row + 1):
        source_id = clean(source_sheet.cell(source_row, 1).value)
        question = clean(source_sheet.cell(source_row, 2).value)
        if not source_id or not question:
            continue

        options = [clean(source_sheet.cell(source_row, col).value) for col in range(3, 7)]
        correct_numbers = {
            int(token) for token in re.findall(r"[1-4]", clean(source_sheet.cell(source_row, 7).value))
        }
        if not correct_numbers:
            raise RuntimeError(f"{source_path.name}, row {source_row}: no correct answer")

        question_count += 1
        written_for_question = 0
        for option_number, answer in enumerate(options, start=1):
            if not answer:
                continue
            output_sheet.append(
                [
                    question_count,
                    question if written_for_question == 0 else "",
                    "",
                    "",
                    answer,
                    "",
                    1 if option_number in correct_numbers else 0,
                ]
            )
            written_for_question += 1
            answer_count += 1

        if written_for_question == 0:
            raise RuntimeError(f"{source_path.name}, row {source_row}: no answer text")
        missing_correct = [n for n in correct_numbers if not options[n - 1]]
        if missing_correct:
            raise RuntimeError(
                f"{source_path.name}, row {source_row}: correct option is blank: {missing_correct}"
            )

    for cell in output_sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    output_sheet.freeze_panes = "A2"
    output_sheet.auto_filter.ref = output_sheet.dimensions
    output_sheet.column_dimensions["A"].width = 13
    output_sheet.column_dimensions["B"].width = 70
    output_sheet.column_dimensions["C"].width = 20
    output_sheet.column_dimensions["D"].width = 12
    output_sheet.column_dimensions["E"].width = 55
    output_sheet.column_dimensions["F"].width = 20
    output_sheet.column_dimensions["G"].width = 12
    for row in output_sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    output_path = OUTPUT / source_path.name
    output_book.save(output_path)
    summary.append((source_path.name, question_count, answer_count))

readme = OUTPUT / "HUONG_DAN.txt"
readme.write_text(
    "Các file Excel trong thư mục này đã được chuyển sang định dạng upload của chức năng 307.\n"
    "Mỗi file có sheet tên EXCEL và 7 cột đúng thứ tự chương trình yêu cầu.\n"
    "Chọn từng file tại chức năng quản lý câu hỏi của 307 để upload.\n"
    "File nguồn ban đầu không bị thay đổi.\n",
    encoding="utf-8-sig",
)

print(f"OUTPUT={OUTPUT}")
print(f"FILES={len(summary)}")
print(f"QUESTIONS={sum(item[1] for item in summary)}")
print(f"ANSWERS={sum(item[2] for item in summary)}")
for name, questions, answers in summary:
    print(f"{name}\t{questions}\t{answers}")
