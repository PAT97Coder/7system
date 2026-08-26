from pathlib import Path
import json
import openpyxl

SOURCE = Path(r"E:\01. Softwares Programming\24. Knowledge System\03. Documents\307\cau hoi chuyen nganh")

for path in sorted(SOURCE.glob("*.xlsx")):
    print(f"\n### {path.name}")
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as exc:
        print("ERROR", repr(exc))
        continue
    for ws in wb.worksheets:
        print(f"SHEET {ws.title!r} rows={ws.max_row} cols={ws.max_column} images={len(ws._images)} merged={len(ws.merged_cells.ranges)}")
        shown = 0
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 30), values_only=True):
            vals = [str(v).replace("\n", "\\n")[:100] if v is not None else "" for v in row]
            if any(vals):
                print(json.dumps(vals[:15], ensure_ascii=False))
                shown += 1
                if shown >= 12:
                    break
